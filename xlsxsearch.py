import sys
try:
    import openpyxl
except:
    sys.path.insert(0, 'c:\\sync\\sync\\hg\\openpyxl')
    import openpyxl
rich_text_available = False
try:
    from openpyxl.cell.rich_text import TextBlock, CellRichText, CellRichTextStr
    from openpyxl.cell.text import InlineFont
    rich_text_available = True
except:
    pass
from pathlib import Path
import os
from copy import copy
import PySimpleGUI as sg
import pprint

from squery import SQuery

col_width = []
pp = pprint.PrettyPrinter(indent=4)

def gen_file_list(folder):
    files = []
    p = Path(folder)
    for i in p.glob('**/*.xlsx'):
        if os.fspath(i).find("xlsxsearch_") < 0:
            files.append(os.fspath(i))
    return files

def run_search(files, squery, query_func, dest_filename, fix_search, w, status):
    status[0]('')
    status[1]('')
    status[2]('')
    w.Refresh()
    #
    wb_w = openpyxl.Workbook()
    ws_w = wb_w.active
    ws_w.title = squery.query
    #
    first = True
    dest_row = 2
    for filename in files:
        wb_r = openpyxl.load_workbook(filename)
        # Get the active worksheet
        ws_r = wb_r.active
        max_row = ws_r.max_row
        max_column = ws_r.max_column
        if len(col_width) < max_column:
            col_width.extend([0] * (max_column - len(col_width)))
        if first:
            status[0]("Took formatting from: " + filename)
            w.Refresh()
            #ws_w.column_dimensions = copy(ws_r.column_dimensions)
            ws_w.sheet_view.rightToLeft = copy(ws_r.sheet_view.rightToLeft)
            ws_w.sheet_format = copy(ws_r.sheet_format) # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.dimensions.html?highlight=SheetFormatProperties#openpyxl.worksheet.dimensions.SheetFormatProperties
            #ws_w.dimensions = copy(ws_r.dimensions)
            ws_w.sheet_properties = copy(ws_r.sheet_properties) #https://foss.heptapod.net/openpyxl/openpyxl/-/blob/branch/3.0/openpyxl/worksheet/properties.py #WorksheetProperties
            #ws_w.merged_cells = copy(ws_r.merged_cells)
            ws_w.page_margins = copy(ws_r.page_margins)
            # page_margins is broken into 3 lines in xlsxwriter
            ws_w.page_setup = copy(ws_r.page_setup) # https://xlsxwriter.readthedocs.io/page_setup.html and check https://foss.heptapod.net/openpyxl/openpyxl/-/blob/branch/3.0/openpyxl/worksheet/page.py
            ws_w.print_options = copy(ws_r.print_options)
            openpyxl_copy_row(ws_r, ws_w, 1, 1)
            first = False
        # ws_r.rows[1:] means we'll skip the first row (the header row).
        for row in range(1, max_row+1):
            found = False
            rich_col = [None]
            for column in range(1, max_column+1):
                cell = ws_r.cell(row=row, column=column)
                c = str(cell.value)
                if c and query_func(c):
                    found = True
                    if fix_search:
                        rstr = []
                        pos = 0
                        for loc, length in sorted(zip(squery.loc, squery.len)):
                            if loc < 0:
                                continue
                            if loc > pos:
                                rstr.append(c[pos:loc])
                            pos = loc + length
                            rstr.append(TextBlock(InlineFont(b=True), c[loc:pos]))
                            
                        if pos < len(c):
                            rstr.append(c[pos:len(c)])
                        new_value = CellRichText(rstr)
                    else:
                        new_value = c
                    rich_col.append(new_value)
                else:
                    rich_col.append(cell.value)

            if found:
                status[1]("Found " + str(dest_row - 1) + (" row" if dest_row == 2 else " rows"))
                w.Refresh()
                openpyxl_copy_row(ws_r, ws_w, row, dest_row, rich_col)
                dest_row += 1
        dims = {}
        for row in ws_w.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for column in range(len(col_width)):
            if col_width[column] > 0:
                ws_w.column_dimensions[openpyxl.utils.get_column_letter(column+1)].width = col_width[column]
    wb_w.save(dest_filename)
    status[2]("Done!")

def openpyxl_copy_row(ws_s, ws_d, src_row, dest_row, rich_col=None):
    max_column = ws_s.max_column
    ws_d.row_dimensions[dest_row] = copy(ws_s.row_dimensions[src_row])
    for column in range(1, max_column + 1):
        cs = ws_s.cell(row=src_row, column=column)
        if rich_col == None:
            new_value = cs.value
        else:
            new_value = rich_col[column]
        if cs.value:
            cd = ws_d.cell(row=dest_row, column=column, value=new_value)
            #print(ws_s.column_dimensions[column].width)
            col_width[column-1] = max(col_width[column-1], ws_s.column_dimensions[openpyxl.utils.get_column_letter(column)].width)
            if cs.has_style:
                cd.font = copy(cs.font)
                cd.border = copy(cs.border)
                cd.fill = copy(cs.fill)
                cd.number_format = copy(cs.number_format)
                cd.protection = copy(cs.protection)
                cd.alignment = copy(cs.alignment)
            if cs.hyperlink:
                cd.hyperlink = copy(cs.hyperlink)
            if cs.comment:
                cd.comment = copy(cs.comment)


#
# start
#
sfb = sg.FolderBrowse()
dfb = sg.FolderBrowse()
status = [None] * 3
for i in range(3):
    status[i] = sg.Text('', size=(100, 1))
layout = [
    [sg.Text('xlsxsearch (23-Nov-2021, $Id$) - Copyright 2020-2021 by Udi Finkelstein', size=(100, 1))],
    [sg.Text('https://github.com/udif/xlsxsearch/', size=(100, 1))],
    [sg.Text('', size=(100, 1))],
    [sg.Text('Source Folders with XLSX files', size=(30, 1)),
     sg.InputText(key='-SFOLDER-', size=(100, 1),
                  default_text = sfb.InitialFolder), sfb],
    [sg.Text('Destination for search result XLSX files', size=(30, 1)),
     sg.InputText(key='-DFOLDER-', size=(100, 1),
                  default_text = dfb.InitialFolder), dfb],
    [sg.Text('Search query', size=(30, 1)), sg.InputText('', size=(100, 1), key='-QUERY-')]
]
if rich_text_available:
    layout.append(
        [sg.Checkbox('Mark search keywords in results in bold style', key='-CB-', size=(40,1), default=True)]
    )
layout.extend([
    [sg.Checkbox('optional "and"', key='-IA-', size=(20,1), default=True, tooltip='If selected, then multiple words without AND or OR are assumed to have implicit AND operator between them')],
    [sg.Text('', size=(100, 1))],
    [sg.Submit()],
    [sg.Text('', size=(100, 1))],
    [status[0]],
    [status[1]],
    [status[2]]
])

w = sg.Window("xlsxsearch", layout, margins=(80, 50))

folder = ''
sq = SQuery()
while True:
    folder = "."
    file_list = []
    event, values = w.read()
    # End program if user closes window or
    # presses the OK button
    if event == "Submit":
        fix_search = rich_text_available and values["-CB-"]
        (o, a, n) = sq.get_keywords()
        a = list(filter(lambda n: len(n) > 0, a))
        if values["-IA-"]:
            a.append('')
        sq.set_keywords(o, a, n)
        sfolder = values["-SFOLDER-"]
        dfolder = values["-DFOLDER-"]
        file_list = gen_file_list(sfolder)
        searchstring = values["-QUERY-"]
        query_func = sq.compile(searchstring)
        if query_func is None:
            status[2]("Illegal query!")
            continue
        filename = searchstring.replace('"', "'")
        dest_file = os.fspath(Path(dfolder).joinpath("xlsxsearch_" + filename + ".xlsx"))
        try:
            run_search(file_list, sq, query_func, dest_file, fix_search, w, status)
        except OSError as err:
            status[2]('OS Error: {}'.format(err))
            w.Refresh()

    if event == sg.WIN_CLOSED:
        break
    if event == sg.WIN_CLOSED:
        w.close()
        sys.exit(0)
    if event == "-SFOLDER-":
        pass
w.close()

#print(files)
